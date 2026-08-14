"""
Ingestion: raw artifact files -> a uniform `IngestedArtifact` the extractors
consume. One loader per format; all deterministic (no model runs here).

Two things every loader must produce:

- `text` — a linearized rendering for the text extractors. For the scanned fax
  this is intentionally empty: `pdfplumber` finds no text layer, which is the
  signal that forces the OCR branch. That branch doesn't shell out to a local
  OCR engine — it flags `needs_ocr` and hands the raw PDF bytes to a
  vision-capable model (see `raw_b64`), which is both simpler and stronger than
  tesseract for degraded scans.
- `blocks` — located sub-units (`Sheet1!C14`, `page 2`, `line 37`, an MTEXT
  position). Every extracted value cites a locator, so the reviewer UI can jump
  to where a value came from. A value with no evidence is a bug.
"""

from __future__ import annotations

import base64
import hashlib
from dataclasses import dataclass, field
from email import message_from_bytes, policy
from email.message import EmailMessage
from pathlib import Path


# --------------------------------------------------------------------------
# Output shape
# --------------------------------------------------------------------------

@dataclass
class Block:
    """A located unit of source content — the anchor for an Evidence locator."""
    locator: str
    text: str


@dataclass
class IngestedArtifact:
    artifact_id: str
    kind: str                       # email|pdf_text|pdf_scanned|xlsx|dxf|transcript
    filename: str
    sha256: str
    bytes: int
    text: str = ""                  # linearized for text extractors ("" forces OCR)
    blocks: list[Block] = field(default_factory=list)
    headers: dict[str, str] = field(default_factory=dict)   # email only
    needs_ocr: bool = False
    page_count: int | None = None
    raw_b64: str | None = None      # base64 of the raw file, for vision extraction

    @property
    def media_type(self) -> str | None:
        return {"pdf_scanned": "application/pdf"}.get(self.kind)


# --------------------------------------------------------------------------
# Dispatch
# --------------------------------------------------------------------------

def ingest_file(path: str | Path, kind: str, artifact_id: str | None = None) -> IngestedArtifact:
    path = Path(path)
    data = path.read_bytes()
    base = IngestedArtifact(
        artifact_id=artifact_id or path.name,
        kind=kind,
        filename=path.name,
        sha256=hashlib.sha256(data).hexdigest()[:16],
        bytes=len(data),
    )
    loader = _LOADERS.get(kind)
    if loader is None:
        raise ValueError(f"no ingester for kind {kind!r}")
    loader(base, path, data)
    return base


# --------------------------------------------------------------------------
# Email
# --------------------------------------------------------------------------

def _load_email(art: IngestedArtifact, path: Path, data: bytes) -> None:
    msg: EmailMessage = message_from_bytes(data, policy=policy.default)
    for h in ("From", "To", "Subject", "Date"):
        if msg[h]:
            art.headers[h] = str(msg[h])

    body = _email_body(msg)
    for h, v in art.headers.items():
        art.blocks.append(Block(locator=f"{h} header", text=v))
    for i, line in enumerate(body.splitlines(), start=1):
        if line.strip():
            art.blocks.append(Block(locator=f"body line {i}", text=line))

    header_txt = "\n".join(f"{h}: {v}" for h, v in art.headers.items())
    art.text = f"{header_txt}\n\n{body}".strip()


def _email_body(msg: EmailMessage) -> str:
    """First text/plain part; the corpus emails are single-part plain text."""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain" and not part.get_filename():
                return part.get_content()
        return ""
    return msg.get_content()


# --------------------------------------------------------------------------
# PDF — text layer
# --------------------------------------------------------------------------

def _load_pdf_text(art: IngestedArtifact, path: Path, data: bytes) -> None:
    import pdfplumber

    pages: list[str] = []
    with pdfplumber.open(path) as pdf:
        art.page_count = len(pdf.pages)
        for pageno, page in enumerate(pdf.pages, start=1):
            txt = page.extract_text() or ""
            pages.append(txt)
            for i, line in enumerate(txt.splitlines(), start=1):
                if line.strip():
                    art.blocks.append(Block(locator=f"page {pageno} line {i}", text=line))
    art.text = "\n".join(pages).strip()


# --------------------------------------------------------------------------
# PDF — scanned (no text layer -> OCR branch)
# --------------------------------------------------------------------------

def _load_pdf_scanned(art: IngestedArtifact, path: Path, data: bytes) -> None:
    import pdfplumber

    with pdfplumber.open(path) as pdf:
        art.page_count = len(pdf.pages)
        text = "\n".join((p.extract_text() or "") for p in pdf.pages).strip()

    if text:
        # A "scanned" artifact that unexpectedly has a text layer — use it and
        # skip OCR. Cheaper and more accurate than sending it to a vision model.
        art.text = text
        for i, line in enumerate(text.splitlines(), start=1):
            if line.strip():
                art.blocks.append(Block(locator=f"page 1 line {i}", text=line))
    else:
        # The real case: empty text layer. Flag OCR and carry the raw bytes for
        # the vision extractor. No local OCR engine involved.
        art.needs_ocr = True
        art.text = ""
        art.raw_b64 = base64.standard_b64encode(data).decode("ascii")


# --------------------------------------------------------------------------
# Spreadsheet
# --------------------------------------------------------------------------

def _load_xlsx(art: IngestedArtifact, path: Path, data: bytes) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    sheet = ws.title

    rows_txt: list[str] = []
    for r, row in enumerate(ws.iter_rows(), start=1):
        cells: list[str] = []
        for c, cell in enumerate(row, start=1):
            if cell.value is None:
                continue
            ref = f"{sheet}!{get_column_letter(c)}{r}"
            val = str(cell.value)
            art.blocks.append(Block(locator=ref, text=val))
            cells.append(f"{get_column_letter(c)}: {val}")
        if cells:
            rows_txt.append(f"row {r} | " + " | ".join(cells))
    wb.close()
    art.text = "\n".join(rows_txt).strip()


# --------------------------------------------------------------------------
# CAD (DXF) — annotation layer only, never geometry
# --------------------------------------------------------------------------

def _load_dxf(art: IngestedArtifact, path: Path, data: bytes) -> None:
    import ezdxf

    doc = ezdxf.readfile(str(path))
    msp = doc.modelspace()
    parts: list[str] = []

    for e in msp:
        dtype = e.dxftype()
        if dtype in ("MTEXT", "TEXT"):
            txt = e.text if dtype == "MTEXT" else e.dxf.text
            txt = (txt or "").strip()
            if not txt:
                continue
            try:
                ins = e.dxf.insert
                loc = f"{dtype}@({ins.x:.0f},{ins.y:.0f})"
            except Exception:
                loc = dtype
            art.blocks.append(Block(locator=loc, text=txt))
            parts.append(txt)
        elif dtype == "DIMENSION":
            label = _dim_label(e)
            if label:
                art.blocks.append(Block(locator="DIMENSION", text=label))
                parts.append(f"DIM {label}")

    art.text = "\n".join(parts).strip()


def _dim_label(e) -> str:
    """The dimension's text override, or its measured value if no override."""
    try:
        override = (e.dxf.text or "").strip()
    except Exception:
        override = ""
    if override and override not in ("<>", " "):
        return override
    try:
        return f'{e.get_measurement():.2f}'
    except Exception:
        return ""


# --------------------------------------------------------------------------
# Call transcript
# --------------------------------------------------------------------------

def _load_transcript(art: IngestedArtifact, path: Path, data: bytes) -> None:
    text = data.decode("utf-8", errors="replace")
    for i, line in enumerate(text.splitlines(), start=1):
        if line.strip():
            art.blocks.append(Block(locator=f"line {i}", text=line))
    art.text = text.strip()


_LOADERS = {
    "email": _load_email,
    "pdf_text": _load_pdf_text,
    "pdf_scanned": _load_pdf_scanned,
    "xlsx": _load_xlsx,
    "dxf": _load_dxf,
    "transcript": _load_transcript,
}
