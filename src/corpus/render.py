"""
Renderers: LeadSpec -> actual files on disk.

Deterministic. Seeded noise only, so a regenerated corpus is byte-comparable and
eval numbers stay reproducible across runs.
"""

from __future__ import annotations

import random
import zlib
from email.message import EmailMessage
from pathlib import Path

from PIL import Image, ImageDraw, ImageFilter, ImageFont

SEED = 20260814


# ---------------------------------------------------------------- email

def render_email(payload: dict, out: Path, attach_dir: Path) -> Path:
    msg = EmailMessage()
    msg["Subject"] = payload["subject"]
    msg["From"] = f'{payload["from_name"]} <{payload["from_email"]}>'
    msg["To"] = "sales@northwind-furniture.example"
    msg["Date"] = payload.get("date", "Thu, 6 Aug 2026 09:14:22 -0400")
    msg["Message-ID"] = f"<{out.stem}@mail.example>"
    msg.set_content(payload["body"])

    for fname in payload.get("attachments", []):
        path = attach_dir / fname
        if not path.exists():
            continue  # attachment rendered later in the pass; see generate.py ordering
        data = path.read_bytes()
        subtype = {
            ".pdf": ("application", "pdf"),
            ".xlsx": ("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            ".dxf": ("application", "dxf"),
        }.get(path.suffix, ("application", "octet-stream"))
        msg.add_attachment(data, maintype=subtype[0], subtype=subtype[1], filename=fname)

    out.write_bytes(msg.as_bytes())
    return out


# ---------------------------------------------------------------- text PDF

def render_pdf_text(payload: dict, out: Path) -> Path:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(out), pagesize=letter)
    width, height = letter
    y = height - 72

    c.setFont("Helvetica-Bold", 13)
    c.drawString(72, y, payload["title"])
    y -= 26

    c.setFont("Helvetica", 10)
    for line in payload["lines"]:
        if y < 72:
            c.showPage()
            c.setFont("Helvetica", 10)
            y = height - 72
        c.drawString(72, y, line)
        y -= 14

    c.save()
    return out


# ---------------------------------------------------------------- scanned PDF

def _load_font(size: int):
    """Find a scalable TrueType face across distros.

    Distro font paths differ (Debian nests under truetype/, Arch does not), so
    we try known locations and then glob before falling back. The fallback is a
    bitmap font that renders the scanned-fax fixture nearly illegible, so it is a
    genuine last resort, not a normal path.
    """
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/gnu-free/FreeSans.ttf",
    ]
    globbed = sorted(Path("/usr/share/fonts").rglob("DejaVuSans.ttf")) + \
              sorted(Path("/usr/share/fonts").rglob("LiberationSans-Regular.ttf")) + \
              sorted(Path("/usr/share/fonts").rglob("*Sans*.ttf"))
    for p in candidates + [str(g) for g in globbed]:
        if Path(p).exists():
            return ImageFont.truetype(p, size)
    return ImageFont.load_default()


def render_pdf_scanned(payload: dict, out: Path) -> Path:
    """No text layer. pdfplumber returns '' and the pipeline must fall to OCR.

    This is the single most useful artifact in the corpus — it is the one that
    proves the router actually inspects the document instead of assuming.
    """
    # zlib.crc32 (not builtin hash(), which is per-process randomized) keeps the
    # scanned image byte-reproducible across runs — the corpus and its recorded
    # cache stay stable.
    rng = random.Random(SEED + zlib.crc32(out.stem.encode()))
    W, H = 1700, 2200          # ~200 dpi letter
    img = Image.new("L", (W, H), 255)
    d = ImageDraw.Draw(img)

    title_font = _load_font(46)
    body_font = _load_font(34)

    y = 180
    d.text((150, y), payload["title"], font=title_font, fill=25)
    y += 90

    for line in payload["lines"]:
        # per-line jitter mimics a real fax feed
        x = 150 + rng.randint(-3, 3)
        d.text((x, y), line, font=body_font, fill=rng.randint(20, 65))
        y += 52

    # speckle
    noise = payload.get("noise", 0.05)
    px = img.load()
    for _ in range(int(W * H * noise * 0.02)):
        px[rng.randrange(W), rng.randrange(H)] = rng.randint(0, 90)

    img = img.filter(ImageFilter.GaussianBlur(0.6))
    img = img.rotate(payload.get("skew_deg", 1.0), resample=Image.BICUBIC,
                     fillcolor=255, expand=False)

    img.convert("RGB").save(str(out), "PDF", resolution=200.0)
    return out


# ---------------------------------------------------------------- xlsx

def render_xlsx(payload: dict, out: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = payload.get("sheet", "Sheet1")

    for r, row in enumerate(payload["rows"], start=1):
        for c, val in enumerate(row, start=1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    for rng_ref in payload.get("merges", []):
        ws.merge_cells(rng_ref)
        anchor = ws[rng_ref.split(":")[0]]
        anchor.alignment = Alignment(horizontal="center")
        anchor.font = Font(bold=True, name="Arial")

    for col, w in zip("ABCDEF", (8, 42, 12, 18, 22, 12)):
        ws.column_dimensions[col].width = w

    wb.save(str(out))
    return out


# ---------------------------------------------------------------- dxf

def render_dxf(payload: dict, out: Path) -> Path:
    """Text + dimension entities only.

    Deliberately no solid geometry: the extraction strategy is to read the
    annotation layer, which is where fabricators actually put the numbers. Do
    not let this creep into geometry parsing — that is a week of work for
    marginal recall.
    """
    import ezdxf

    doc = ezdxf.new("R2010", setup=True)
    msp = doc.modelspace()

    for x, y, text in payload.get("texts", []):
        msp.add_mtext(text, dxfattribs={"style": "OpenSans", "char_height": 4.0}) \
           .set_location((x, y))

    for dim in payload.get("dims", []):
        d = msp.add_aligned_dim(
            p1=dim["p1"], p2=dim["p2"],
            distance=dim.get("distance", 10),
            text=dim.get("label", "<>"),
        )
        d.render()

    doc.saveas(str(out))
    return out


# ---------------------------------------------------------------- transcript

def render_transcript(payload: dict, out: Path) -> Path:
    body = [payload.get("meta", ""), ""] + payload["lines"]
    out.write_text("\n".join(body) + "\n", encoding="utf-8")
    return out


RENDERERS = {
    "email": render_email,
    "pdf_text": render_pdf_text,
    "pdf_scanned": render_pdf_scanned,
    "xlsx": render_xlsx,
    "dxf": render_dxf,
    "transcript": render_transcript,
}
